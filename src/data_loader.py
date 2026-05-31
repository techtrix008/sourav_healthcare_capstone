from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .models import PatientRecord


def _normalize_phone(value: Any) -> str | None:
    if value is None:
        return None
    return str(value).strip()


def _patient_id(name: str, phone: str | None, source: str, row_number: int) -> str:
    slug = "-".join(name.lower().split())
    suffix = phone or f"{source}-{row_number}"
    suffix = "".join(ch for ch in str(suffix) if ch.isalnum())
    return f"{slug}-{suffix[-6:]}"


def load_excel_records(path: Path) -> list[PatientRecord]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(header).strip() for header in rows[0]]
    records: list[PatientRecord] = []
    seen: set[tuple[str, str | None, str | None]] = set()

    for row_number, row in enumerate(rows[1:], start=2):
        item = dict(zip(headers, row))
        name = str(item.get("Name") or "").strip()
        if not name:
            continue
        phone = _normalize_phone(item.get("Phone_number"))
        summary = item.get("Summary")
        dedupe_key = (name.lower(), phone, str(summary or ""))
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)

        records.append(
            PatientRecord(
                patient_id=_patient_id(name, phone, path.name, row_number),
                name=name,
                age=int(item["Age"]) if item.get("Age") is not None else None,
                gender=item.get("Gender"),
                phone=phone,
                email=item.get("Email"),
                address=item.get("Address"),
                summary=str(summary).strip() if summary else None,
                source=path.name,
            )
        )
    return records
